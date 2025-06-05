import openpyxl
from openpyxl.styles import Font, Alignment
import os

# a raw string so back-slashes donâ€™t get treated as escapes
file_name = r"C:\Education\Project\Student_mark_sys.xlsx"

def excel():
        if not os.path.exists(file_name):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Student Marks"
        else:
            wb = openpyxl.load_workbook(file_name)

        if "Student Marks" in wb.sheetnames:
            ws = wb["Student Marks"]
        else:
            ws = wb.active
            ws.title = "Student Marks"

        if ws.cell(row=1, column=1).value != "Student Marks Management System":
            ws.merge_cells("A1:K1")
            ws["A1"] = "Student Marks Management System"
            ws["A1"].font = Font(size=20, bold=True)
            ws["A1"].alignment = Alignment(horizontal="center")

        expected_headers = [
            "Roll No", "Name", "Maths", "OS", "English",
            "Database Fundamentals", "Programming Skills",
            "Total", "Percentage", "Grade", "Pass/Fail"
        ]
        if ws.cell(row=2, column=1).value != expected_headers[0]:
            for col_num, header in enumerate(expected_headers, start=1):
                ws.cell(row=2, column=col_num, value=header)

        ws.freeze_panes = "A3"
        wb.save(file_name)


def open_excel(wb, file_name):
    try:
        wb.save(file_name)
    except PermissionError:
        print("Permission denied: Please make sure the Excel file is closed.")


def calculation(marks_list, passing=33):
    if any(mark < passing for mark in marks_list):
        total = sum(marks_list)
        percentage = total / len(marks_list)
        return total, percentage, "-", "Fail"

    total = sum(marks_list)
    percentage = total / len(marks_list)

    if percentage >= 90:
        grade = "A+"
    elif percentage >= 80:
        grade = "A"
    elif percentage >= 70:
        grade = "B"
    elif percentage >= 60:
        grade = "C"
    elif percentage >= 50:
        grade = "D"
    else:
        grade = "F"

    return total, percentage, grade, "Pass"

def add_student(roll_no, name, maths, os_mark, eng, db_fund, prog_skill):
    try:
        marks = list(map(int, [maths, os_mark, eng, db_fund, prog_skill]))

        if any(mark < 0 or mark > 100 for mark in marks):
            print("All subject marks must be between 0 and 100.")
            return

        wb = openpyxl.load_workbook(file_name)
        if "Student Marks" in wb.sheetnames:
            ws = wb["Student Marks"]
        else:
            ws = wb.active
            ws.title = "Student Marks"


        # Duplicate-roll-number check
        for row in ws.iter_rows(min_row=3, max_col=1, values_only=True):
            if str(row[0]) == roll_no:
                print(f"Roll number {roll_no} already exists. Please use a unique roll number.")
                return

        total, percentage, grade, pass_fail = calculation(marks)
        ws.append([roll_no, name, *marks, total, round(percentage, 2), grade, pass_fail])
        wb.save(file_name)
        print(f"Student {name} added successfully with grade '{grade}' ({pass_fail}).")

    except ValueError:
        print("Please enter valid integer marks.")

def update_student_marks(roll_no):
    try:
        wb = openpyxl.load_workbook(file_name)
        ws = wb["Student Marks"]

        for row in range(3, ws.max_row + 1):
            cell_roll = ws.cell(row=row, column=1).value
            if str(cell_roll) == roll_no:
                print(f"Updating marks for Roll No: {roll_no}")

                # Prompt only for marks, allow blank to keep current
                maths = input("Enter new Maths marks (leave blank to keep current): ")
                os_mark = input("Enter new OS marks (leave blank to keep current): ")
                eng = input("Enter new English marks (leave blank to keep current): ")
                db_fund = input("Enter new Database Fundamentals marks (leave blank to keep current): ")
                prog_skill = input("Enter new Programming Skills marks (leave blank to keep current): ")

                current_marks = [
                    ws.cell(row=row, column=3).value,
                    ws.cell(row=row, column=4).value,
                    ws.cell(row=row, column=5).value,
                    ws.cell(row=row, column=6).value,
                    ws.cell(row=row, column=7).value,
                ]

                try:
                    maths = int(maths) if maths.strip() else current_marks[0]
                    os_mark = int(os_mark) if os_mark.strip() else current_marks[1]
                    eng = int(eng) if eng.strip() else current_marks[2]
                    db_fund = int(db_fund) if db_fund.strip() else current_marks[3]
                    prog_skill = int(prog_skill) if prog_skill.strip() else current_marks[4]
                except ValueError:
                    print("Invalid marks entered. Update cancelled.")
                    return

                marks = [maths, os_mark, eng, db_fund, prog_skill]

                if any(mark < 0 or mark > 100 for mark in marks):
                    print("All marks must be between 0 and 100. Update cancelled.")
                    return

                # Update marks in sheet
                for i, mark in enumerate(marks, start=3):
                    ws.cell(row=row, column=i, value=mark)

                # Recalculate totals, percentage, grade, pass/fail
                total, percentage, grade, pass_fail = calculation(marks)
                ws.cell(row=row, column=8, value=total)
                ws.cell(row=row, column=9, value=round(percentage, 2))
                ws.cell(row=row, column=10, value=grade)
                ws.cell(row=row, column=11, value=pass_fail)

                wb.save(file_name)
                print(f"Marks updated successfully for Roll No {roll_no}.")
                return

        print(f"Roll number {roll_no} not found.")

    except PermissionError:
        print("Permission denied: Please close the Excel file before updating marks.")

def display_students():
    wb = openpyxl.load_workbook(file_name)
    ws = wb["Student Marks"]
    if ws.max_row <= 2:
        print("No students found.")
        return
    print("\nStudent Details:")
    # start at row 3 to skip title + headers
    for row in ws.iter_rows(min_row=3, values_only=True):
        print(row)


def main():
    try:
        excel()  # Initialize or verify Excel file
    except PermissionError:
        print(f"Permission denied: Please close '{file_name}' if it is open in Excel.")
        return 

    print("Welcome to the Student Marks Management System")

    while True:
        print("\nStudent Marks System")
        print("1. Add Student")
        print("2. Update Student Marks")
        print("3. Display All Students")
        print("4. Exit")

        choice = input("Enter your choice: ")

        if choice == "1":
            roll_no = input("Enter roll number: ")
            name = input("Enter name: ")
            maths = input("Enter Maths marks: ")
            os_mark = input("Enter OS marks: ")
            eng = input("Enter English marks: ")
            db_fund = input("Enter Database Fundamentals marks: ")
            prog_skill = input("Enter Programming Skills marks: ")
            add_student(roll_no, name, maths, os_mark, eng, db_fund, prog_skill)

        elif choice == "2":
            roll_no = input("Enter roll number to update marks: ")
            update_student_marks(roll_no)

        elif choice == "3":
            display_students()

        elif choice == "4":
            print("Exiting program.")
            break

        else:
            print("Invalid choice. Please try again.")
            
if __name__ == "__main__":
    main()
    
